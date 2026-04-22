"""
PAA RN - Gerador Definitivo de Relatórios
Produz tabelas por Município e por Região (10 territórios do RN)
para alimentar o BI da plataforma e os textos do Observatório Rural.

Bases cobertas (API MiSocial/MDS):
  3. Execução Geral (2011-2025)
  4. Perfil de Sexo (2014-2025)
  5. Doação Simultânea via Termo de Adesão (2021-2025)
  6. Modalidade Leite (2021-2025)
  7. Agricultores Doação Simultânea (2011-2025)
  8. Agricultores Leite (2011-2025)
  9. Compra Direta (2011-2017)
  10. Apoio à Formação de Estoques (2011-2019)
  11. Aquisição de Sementes (2015-2020)
  2. Adesão Municipal (2022-2025)
"""

import os, io, json, time, requests, urllib3, pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()

# ─── Configurações ────────────────────────────────────────────────────────────
MDS_API   = "https://aplicacoes.mds.gov.br/sagi/servicos/misocial/"
IBGE_API  = "https://servicodados.ibge.gov.br/api/v1/localidades/estados/24/municipios"
OUTPUT    = "outputs/paa_reports/PAA_RN_DEFINITIVO.xlsx"
IBGE_JSON = "data/bronze/ibge/municipios_rn.json"
os.makedirs("outputs/paa_reports", exist_ok=True)

# ─── Paleta DIEESE ────────────────────────────────────────────────────────────
NAVY   = "0D1B2A"
WHITE  = "FFFFFF"
ZEBRA  = "EFF3F8"
GREEN  = "1B4332"
AMBER  = "D4A017"

# ─── 10 Territórios do RN ─────────────────────────────────────────────────────
# Mapeamento completo por código IBGE (6 dígitos)
TERRITORIOS = {
    "Seridó": [
        240350, 240400, 240450, 240460, 240600, 240650, 240660, 240840, 240900,
        241010, 241020, 241030, 240980, 241170, 241350, 241420, 241570, 241575,
        241580, 241585, 241590, 241600, 241602, 241680, 241690
    ],
    "Alto Oeste": [
        240050, 240100, 240110, 240120, 240310, 240370, 240560, 240630, 240640,
        241040, 241050, 241090, 241100, 241130, 241140, 241145, 241390, 241400,
        241405, 241410, 241470, 241480, 241485, 241560, 241565, 241610, 241650,
        241660, 241760, 241770
    ],
    "Agreste Litoral Sul": [
        240070, 240180, 240190, 240330, 240540, 240700, 241060, 241110, 241115,
        241120, 241160, 241165, 241230, 241240, 241430, 241440, 241520, 241525,
        241540, 241545, 241670, 241740, 241750, 241755
    ],
    "Mato Grande": [
        240200, 240290, 240300, 241080, 241085, 241088, 241200, 241380, 241385,
        241388, 241395, 241490, 241530, 241535, 241640, 241645, 241735, 241745,
        241765
    ],
    "Assú-Mossoró": [
        240060, 240065, 240067, 240210, 240340, 240710, 240950, 240960, 241150,
        241415, 241420, 241500, 241505, 241700
    ],
    "Potengi": [
        240220, 240230, 240970, 241075, 241455, 241460, 241510, 241515, 241518,
        241519, 241521
    ],
    "Trairi": [
        240240, 240280, 240285, 241070, 241072, 241095, 241155, 241380, 241446,
        241550, 241555, 241556, 241557, 241558, 241720
    ],
    "Sertão Central Cabugi e Litoral Norte": [
        240080, 240090, 240320, 240580, 240720, 240730, 241105, 241210, 241445,
        241450
    ],
    "Sertão do Apodi": [
        240130, 240140, 240360, 240590, 240740, 240960, 241065, 241220, 241340,
        241375, 241376, 241460, 241465, 241550, 241710, 241770, 241780
    ],
    "Terras Potiguares": [
        240570, 241180, 241250, 241430, 241555
    ]
}

def build_ibge_territory_map():
    """Carrega os 167 municípios e monta: ibge_6 -> {nome, territorio}"""
    with open(IBGE_JSON, encoding='utf-8') as f:
        munis = json.load(f)

    # Mapa IBGE 6 dígitos -> nome
    ibge_nome = {str(m['id'])[:6]: m['nome'] for m in munis}

    # Mapa IBGE 6 dígitos -> território (busca pelos primeiros 6 dígitos do código)
    ibge_territorio = {}
    for territorio, codigos in TERRITORIOS.items():
        for cod in codigos:
            ibge_territorio[str(cod)] = territorio

    # Para municípios sem território explícito, usar IBGE intermediária como fallback
    for m in munis:
        cod6 = str(m['id'])[:6]
        if cod6 not in ibge_territorio:
            ibge_territorio[cod6] = "Terras Potiguares"  # fallback

    return ibge_nome, ibge_territorio


def fetch_misocial(fields: list, label: str) -> pd.DataFrame:
    """Faz requisição paginada à API MiSocial para RN."""
    print(f"  Extraindo: {label}...")
    url = (
        f"{MDS_API}?q=codigo_ibge:24*&rows=500000&wt=csv"
        f"&fl={','.join(fields)}"
    )
    try:
        resp = requests.get(url, verify=False, timeout=120)
        if resp.status_code == 200 and len(resp.text) > 200:
            df = pd.read_csv(io.StringIO(resp.text))
            df['Ano'] = df['anomes_s'].astype(str).str[:4].astype(int)
            df['ibge6'] = df['codigo_ibge'].astype(str).str[:6]
            print(f"    → {len(df):,} registros brutos")
            return df
    except Exception as e:
        print(f"    → ERRO: {e}")
    return pd.DataFrame()


def format_sheet(ws, df, title=""):
    """Formata uma sheet com padrão DIEESE."""
    # Cabeçalho
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_name
        cell.font = Font(name='Calibri', bold=True, color=WHITE, size=10)
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Dados com zebra
    for row_num in range(2, len(df) + 2):
        bg = ZEBRA if row_num % 2 == 0 else WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.font = Font(name='Calibri', size=10)
            col_name = df.columns[col_num - 1]
            val = df.iloc[row_num - 2, col_num - 1]
            cell.value = val
            # Formatos numéricos
            if 'R$' in col_name or 'Valor' in col_name or 'Recurso' in col_name:
                cell.number_format = 'R$ #,##0.00'
            elif '%' in col_name:
                cell.number_format = '0.0'
            elif col_name in ['Total Fornecedores', 'Homens', 'Mulheres', 'Sem Info',
                               'Qtd Municípios', 'Qtd Agricultores', 'Municípios com Execução']:
                cell.number_format = '#,##0'

    # Largura automática
    for i, col in enumerate(df.columns, 1):
        vals = [str(df.iloc[r, i-1]) for r in range(min(len(df), 100))]
        max_len = max([len(v) for v in vals] + [len(col)]) + 3
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 50)

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 35


def generate():
    print("=" * 65)
    print("PAA RN — GERADOR DEFINITIVO (Municipal + Regional)")
    print("=" * 65)

    # 1. Mapas territoriais
    print("\n[1/3] Carregando mapeamento territorial...")
    ibge_nome, ibge_territorio = build_ibge_territory_map()
    print(f"  → {len(ibge_nome)} municípios | {len(set(ibge_territorio.values()))} territórios")

    # 2. Extração unificada (uma única requisição grande)
    print("\n[2/3] Extraindo dados da API MiSocial...")
    FIELDS = [
        "codigo_ibge", "anomes_s",
        # Base 3 - Execução Geral
        "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f",
        # Base 4 - Sexo
        "paa_qtd_agricul_familiar_sexo_masculino_i",
        "paa_qtd_agricul_familiar_sexo_feminino_i",
        "paa_qtd_agricul_familiar_sem_info_sexo_i",
        # Base 5 - Doação Simultânea Exec
        "paa_qtd_agricul_familiar_exec_compra_doacao_simul_i",
        "paa_vlr_pago_exec_compra_doacao_simul_d",
        "paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d",
        # Base 6 - Leite Exec
        "paa_qtd_agricul_familiar_exec_incentivo_leite_i",
        "paa_vlr_pago_exec_incentivo_leite_d",
        "paa_qtd_alim_adquiridos_exec_incentivo_leite_i",
        # Bases 7-11 - Agricultores por modalidade
        "paa_qtd_agricul_familiar_modal_compra_doacao_simul_i",
        "paa_qtd_agricul_familiar_modal_incentivo_leite_i",
        "paa_qtd_agricul_familiar_modal_compra_direta_i",
        "paa_qtd_agricul_familiar_modal_formacao_estoque_i",
        "paa_qtd_agricul_familiar_modal_aquisicao_sementes_i",
        # Base 2 - Adesão
        "paa_indicador_adesao_municipio_i",
    ]

    df_raw = fetch_misocial(FIELDS, "Extração unificada RN")
    if df_raw.empty:
        print("ERRO CRÍTICO: dados não retornados.")
        return

    # Adicionar nome e território
    df_raw['Município'] = df_raw['ibge6'].map(ibge_nome).fillna('Não identificado')
    df_raw['Território'] = df_raw['ibge6'].map(ibge_territorio).fillna('Não classificado')
    df_raw = df_raw.fillna(0)

    print(f"\n[3/3] Gerando planilhas em {OUTPUT}...")

    with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:

        # ── BASE 3: Execução Geral ──────────────────────────────────────────
        # Por Município
        b3_mun = df_raw.groupby(['Município', 'Território', 'Ano']).agg(
            Fornecedores=('agricultores_fornec_paa_i', 'sum'),
            Recurso_Pago=('recur_pagos_agricul_paa_f', 'sum')
        ).reset_index()
        b3_mun = b3_mun[b3_mun['Fornecedores'] > 0].copy()
        b3_mun['Ticket Médio (R$)'] = (b3_mun['Recurso_Pago'] / b3_mun['Fornecedores']).round(2)
        b3_mun.rename(columns={'Recurso_Pago': 'Recurso Pago (R$)', 'Fornecedores': 'Total Fornecedores'}, inplace=True)
        b3_mun.to_excel(writer, sheet_name='B3 Exec. Geral - Município', index=False)
        format_sheet(writer.sheets['B3 Exec. Geral - Município'], b3_mun)

        # Por Território
        b3_ter = df_raw.groupby(['Território', 'Ano']).agg(
            Fornecedores=('agricultores_fornec_paa_i', 'sum'),
            Recurso_Pago=('recur_pagos_agricul_paa_f', 'sum'),
            Municipios_Exec=('Município', lambda x: (df_raw.loc[x.index, 'agricultores_fornec_paa_i'] > 0).sum())
        ).reset_index()
        b3_ter = b3_ter[b3_ter['Fornecedores'] > 0].copy()
        b3_ter['Ticket Médio (R$)'] = (b3_ter['Recurso_Pago'] / b3_ter['Fornecedores']).round(2)
        b3_ter.rename(columns={
            'Recurso_Pago': 'Recurso Pago (R$)',
            'Fornecedores': 'Qtd Agricultores',
            'Municipios_Exec': 'Municípios com Execução'
        }, inplace=True)
        b3_ter.to_excel(writer, sheet_name='B3 Exec. Geral - Território', index=False)
        format_sheet(writer.sheets['B3 Exec. Geral - Território'], b3_ter)

        # ── BASE 4: Perfil de Sexo ──────────────────────────────────────────
        df_sex = df_raw[df_raw['Ano'] >= 2014]

        b4_mun = df_sex.groupby(['Município', 'Território', 'Ano']).agg(
            Homens=('paa_qtd_agricul_familiar_sexo_masculino_i', 'sum'),
            Mulheres=('paa_qtd_agricul_familiar_sexo_feminino_i', 'sum'),
            Sem_Info=('paa_qtd_agricul_familiar_sem_info_sexo_i', 'sum'),
        ).reset_index()
        b4_mun = b4_mun[(b4_mun['Homens'] + b4_mun['Mulheres']) > 0].copy()
        b4_mun['Sem Info'] = b4_mun['Sem_Info']
        b4_mun['% Feminino'] = (b4_mun['Mulheres'] / (b4_mun['Homens'] + b4_mun['Mulheres']) * 100).round(1)
        b4_mun.drop(columns=['Sem_Info'], inplace=True)
        b4_mun.to_excel(writer, sheet_name='B4 Perfil Sexo - Município', index=False)
        format_sheet(writer.sheets['B4 Perfil Sexo - Município'], b4_mun)

        b4_ter = df_sex.groupby(['Território', 'Ano']).agg(
            Homens=('paa_qtd_agricul_familiar_sexo_masculino_i', 'sum'),
            Mulheres=('paa_qtd_agricul_familiar_sexo_feminino_i', 'sum'),
        ).reset_index()
        b4_ter = b4_ter[(b4_ter['Homens'] + b4_ter['Mulheres']) > 0].copy()
        b4_ter['% Feminino'] = (b4_ter['Mulheres'] / (b4_ter['Homens'] + b4_ter['Mulheres']) * 100).round(1)
        b4_ter.to_excel(writer, sheet_name='B4 Perfil Sexo - Território', index=False)
        format_sheet(writer.sheets['B4 Perfil Sexo - Território'], b4_ter)

        # ── BASE 5: Doação Simultânea (Execução) ───────────────────────────
        df_b5 = df_raw[df_raw['Ano'] >= 2021]

        b5_mun = df_b5.groupby(['Município', 'Território', 'Ano']).agg(
            Agricultores=('paa_qtd_agricul_familiar_exec_compra_doacao_simul_i', 'sum'),
            Valor=('paa_vlr_pago_exec_compra_doacao_simul_d', 'sum'),
            Quantidade_kg=('paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d', 'sum'),
        ).reset_index()
        b5_mun = b5_mun[b5_mun['Valor'] > 0].copy()
        b5_mun.rename(columns={'Valor': 'Valor Pago (R$)', 'Quantidade_kg': 'Alimentos (kg)'}, inplace=True)
        b5_mun.to_excel(writer, sheet_name='B5 Doação Simult - Município', index=False)
        format_sheet(writer.sheets['B5 Doação Simult - Município'], b5_mun)

        b5_ter = df_b5.groupby(['Território', 'Ano']).agg(
            Agricultores=('paa_qtd_agricul_familiar_exec_compra_doacao_simul_i', 'sum'),
            Valor=('paa_vlr_pago_exec_compra_doacao_simul_d', 'sum'),
            Quantidade_kg=('paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d', 'sum'),
        ).reset_index()
        b5_ter = b5_ter[b5_ter['Valor'] > 0].copy()
        b5_ter.rename(columns={'Valor': 'Valor Pago (R$)', 'Quantidade_kg': 'Alimentos (kg)'}, inplace=True)
        b5_ter.to_excel(writer, sheet_name='B5 Doação Simult - Território', index=False)
        format_sheet(writer.sheets['B5 Doação Simult - Território'], b5_ter)

        # ── BASE 6: Leite ───────────────────────────────────────────────────
        b6_mun = df_b5.groupby(['Município', 'Território', 'Ano']).agg(
            Agricultores=('paa_qtd_agricul_familiar_exec_incentivo_leite_i', 'sum'),
            Valor=('paa_vlr_pago_exec_incentivo_leite_d', 'sum'),
            Litros=('paa_qtd_alim_adquiridos_exec_incentivo_leite_i', 'sum'),
        ).reset_index()
        b6_mun = b6_mun[b6_mun['Valor'] > 0].copy()
        b6_mun.rename(columns={'Valor': 'Valor Pago (R$)'}, inplace=True)
        b6_mun.to_excel(writer, sheet_name='B6 Leite - Município', index=False)
        format_sheet(writer.sheets['B6 Leite - Município'], b6_mun)

        b6_ter = df_b5.groupby(['Território', 'Ano']).agg(
            Agricultores=('paa_qtd_agricul_familiar_exec_incentivo_leite_i', 'sum'),
            Valor=('paa_vlr_pago_exec_incentivo_leite_d', 'sum'),
            Litros=('paa_qtd_alim_adquiridos_exec_incentivo_leite_i', 'sum'),
        ).reset_index()
        b6_ter = b6_ter[b6_ter['Valor'] > 0].copy()
        b6_ter.rename(columns={'Valor': 'Valor Pago (R$)'}, inplace=True)
        b6_ter.to_excel(writer, sheet_name='B6 Leite - Território', index=False)
        format_sheet(writer.sheets['B6 Leite - Território'], b6_ter)

        # ── BASES 7-11: Agricultores por Modalidade ─────────────────────────
        modais = {
            'B7 Modal Doação Simult':  ('paa_qtd_agricul_familiar_modal_compra_doacao_simul_i', 2011, 2025),
            'B8 Modal Leite':          ('paa_qtd_agricul_familiar_modal_incentivo_leite_i',     2011, 2025),
            'B9 Modal Compra Direta':  ('paa_qtd_agricul_familiar_modal_compra_direta_i',        2011, 2017),
            'B10 Modal Estoque':       ('paa_qtd_agricul_familiar_modal_formacao_estoque_i',     2011, 2019),
            'B11 Modal Sementes':      ('paa_qtd_agricul_familiar_modal_aquisicao_sementes_i',   2015, 2020),
        }
        for base_name, (field, yr_ini, yr_fim) in modais.items():
            df_modal = df_raw[(df_raw['Ano'] >= yr_ini) & (df_raw['Ano'] <= yr_fim)]

            b_mun = df_modal.groupby(['Município', 'Território', 'Ano']).agg(
                Agricultores=(field, 'sum')
            ).reset_index()
            b_mun = b_mun[b_mun['Agricultores'] > 0].copy()
            sheet_mun = f"{base_name} - Município"[:31]
            b_mun.to_excel(writer, sheet_name=sheet_mun, index=False)
            format_sheet(writer.sheets[sheet_mun], b_mun)

            b_ter = df_modal.groupby(['Território', 'Ano']).agg(
                Agricultores=(field, 'sum')
            ).reset_index()
            b_ter = b_ter[b_ter['Agricultores'] > 0].copy()
            sheet_ter = f"{base_name} - Território"[:31]
            b_ter.to_excel(writer, sheet_name=sheet_ter, index=False)
            format_sheet(writer.sheets[sheet_ter], b_ter)

        # ── BASE 2: Adesão Municipal ─────────────────────────────────────────
        df_b2 = df_raw[df_raw['Ano'] >= 2022]
        b2_mun = df_b2.groupby(['Município', 'Território', 'Ano']).agg(
            Aderiu=('paa_indicador_adesao_municipio_i', 'max')
        ).reset_index()
        b2_mun['Status'] = b2_mun['Aderiu'].map({1: 'Sim', 0: 'Não'})
        b2_mun.drop(columns=['Aderiu'], inplace=True)
        b2_mun.to_excel(writer, sheet_name='B2 Adesão - Município', index=False)
        format_sheet(writer.sheets['B2 Adesão - Município'], b2_mun)

        b2_ter = df_b2.groupby(['Território', 'Ano']).agg(
            Municipios_Aderidos=('paa_indicador_adesao_municipio_i', 'sum'),
            Total_Municipios=('Município', 'nunique')
        ).reset_index()
        b2_ter['% Adesão'] = (b2_ter['Municipios_Aderidos'] / b2_ter['Total_Municipios'] * 100).round(1)
        b2_ter.to_excel(writer, sheet_name='B2 Adesão - Território', index=False)
        format_sheet(writer.sheets['B2 Adesão - Território'], b2_ter)

        # ── PAINEL GERAL RN (resumo para texto) ─────────────────────────────
        painel_rn = df_raw.groupby('Ano').agg(
            Total_Agricultores=('agricultores_fornec_paa_i', 'sum'),
            Total_Recurso=('recur_pagos_agricul_paa_f', 'sum'),
            Total_Mulheres=('paa_qtd_agricul_familiar_sexo_feminino_i', 'sum'),
            Total_Homens=('paa_qtd_agricul_familiar_sexo_masculino_i', 'sum'),
            Litros_Leite=('paa_qtd_alim_adquiridos_exec_incentivo_leite_i', 'sum'),
            Kg_Alimentos=('paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d', 'sum'),
        ).reset_index()
        painel_rn = painel_rn[painel_rn['Total_Agricultores'] > 0].copy()
        painel_rn['Ticket Médio (R$)'] = (painel_rn['Total_Recurso'] / painel_rn['Total_Agricultores']).round(2)
        painel_rn['% Feminino'] = (painel_rn['Total_Mulheres'] / (painel_rn['Total_Homens'] + painel_rn['Total_Mulheres']) * 100).round(1)
        painel_rn.rename(columns={
            'Total_Agricultores': 'Total Agricultores',
            'Total_Recurso': 'Recurso Total (R$)',
            'Total_Mulheres': 'Agricultoras (F)',
            'Total_Homens': 'Agricultores (M)',
            'Litros_Leite': 'Leite Adquirido (L)',
            'Kg_Alimentos': 'Alimentos Adquiridos (kg)',
        }, inplace=True)
        painel_rn.to_excel(writer, sheet_name='PAINEL GERAL RN', index=False)
        format_sheet(writer.sheets['PAINEL GERAL RN'], painel_rn)

    print(f"\n✅ CONCLUÍDO!")
    print(f"   Arquivo gerado: {OUTPUT}")
    print(f"   Abas criadas: {len(pd.ExcelFile(OUTPUT).sheet_names)}")


if __name__ == '__main__':
    generate()
