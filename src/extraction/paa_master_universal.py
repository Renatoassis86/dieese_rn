import pandas as pd
import requests
import io
import os
import urllib3
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configurações
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
MDS_API_URL = "https://aplicacoes.mds.gov.br/sagi/servicos/misocial/"
OUTPUT_MASTER = "outputs/paa_reports/PAA_OBSERVATORIO_FINAL_MASTER.xlsx"
os.makedirs("outputs/paa_reports", exist_ok=True)

# 1. Mapeamento de Territórios do RN (OficialDIEESE/Seminário)
TERRITORIOS_RN = {
    "Seridó": [240350, 240400, 240450, 240460, 240600, 240650, 240660, 240840, 240900, 241010, 241020, 241030, 240980, 241170, 241350, 241420, 241570, 241575, 241580, 241585, 241590, 241600, 241602, 241680, 241690],
    "Alto Oeste": [240050, 240100, 240110, 240120, 240310, 240370, 240560, 240630, 240640, 241040, 241050, 241090, 241100, 241130, 241140, 241145, 241390, 241400, 241405, 241410, 241470, 241480, 241485, 241560, 241565, 241610, 241650, 241660, 241760, 241770],
    "Agreste Litoral Sul": [240070, 240180, 240190, 240330, 240540, 240700, 241060, 241110, 241115, 241120, 241160, 241165, 241230, 241240, 241430, 241440, 241520, 241525, 241540, 241545, 241670, 241740, 241750, 241755],
    "Mato Grande": [240200, 240290, 240300, 241080, 241085, 241088, 241200, 241380, 241385, 241388, 241395, 241490, 241530, 241535, 241640, 241645, 241735, 241745, 241765],
    "Assú-Mossoró": [240060, 240065, 240067, 240210, 240340, 240710, 240950, 240960, 241150, 241415, 241420, 241500, 241505, 241700],
    "Potengi": [240220, 240230, 240970, 241075, 241455, 241460, 241510, 241515, 241518, 241519, 241521],
    "Trairi": [240240, 240280, 240285, 241070, 241072, 241095, 241155, 241380, 241446, 241550, 241555, 241556, 241557, 241558, 241720],
    "Sertão Central Cabugi e Litoral Norte": [240080, 240090, 240320, 240580, 240720, 240730, 241105, 241210, 241445, 241450],
    "Sertão do Apodi": [240130, 240140, 240360, 240590, 240740, 240960, 241065, 241220, 241340, 241375, 241376, 241460, 241465, 241550, 241710, 241770, 241780],
    "Terras Potiguares": [240570, 241180, 241250, 241430, 241555]
}

UF_NE = {'21': 'MA', '22': 'PI', '23': 'CE', '24': 'RN', '25': 'PB', '26': 'PE', '27': 'AL', '28': 'SE', '29': 'BA'}

FIELDS = [
    "codigo_ibge", "anomes_s", "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f",
    "paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i",
    "paa_vlr_pago_exec_compra_doacao_simul_d", "paa_vlr_pago_exec_incentivo_leite_d",
    "paa_qtd_agricul_familiar_modal_compra_doacao_simul_i", "paa_qtd_agricul_familiar_modal_incentivo_leite_i",
    "paa_qtd_agricul_familiar_modal_compra_direta_i", "paa_qtd_agricul_familiar_modal_formacao_estoque_i"
]

def format_excel(writer, sheet, df):
    ws = writer.sheets[sheet]
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, val in enumerate(df.columns.values):
        cell = ws.cell(row=1, column=col_num + 1)
        cell.fill, cell.font = header_fill, header_font
    for i in range(len(df.columns)): ws.column_dimensions[get_column_letter(i + 1)].width = 20

def fetch_data(query_uf="*"):
    url = f"{MDS_API_URL}?q=codigo_ibge:{query_uf}&rows=2000000&wt=csv&fl={','.join(FIELDS)}"
    resp = requests.get(url, verify=False, timeout=180)
    df = pd.read_csv(io.StringIO(resp.text))
    df['Ano'] = df['anomes_s'].astype(str).str[:4]
    df['UF'] = df['codigo_ibge'].astype(str).str[:2].map(UF_NE)
    for c in FIELDS[2:]: df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    return df

def generate():
    print("🚀 Iniciando Extração Universal Nordeste + RN...")
    
    # 1. Extração Nordeste (Agregado)
    df_ne_full = []
    for cod_uf in UF_NE.keys():
        print(f"Buscando {UF_NE[cod_uf]}...")
        df_ne_full.append(fetch_data(f"{cod_uf}*"))
    
    df_all = pd.concat(df_ne_full)
    
    # --- NIVEL 1: COMPARATIVO NORDESTE ---
    ne_comparativo = df_all.groupby(['UF', 'Ano']).agg({
        "recur_pagos_agricul_paa_f": "sum",
        "agricultores_fornec_paa_i": "sum"
    }).reset_index()
    ne_comparativo.columns = ['Estado', 'Ano', 'Valor Pago (R$)', 'Agricultores']
    
    # --- NIVEL 2: RN GERAL ---
    df_rn = df_all[df_all['UF'] == 'RN'].copy()
    rn_geral = df_rn.groupby('Ano').agg({
        "recur_pagos_agricul_paa_f": "sum",
        "agricultores_fornec_paa_i": "sum",
        "paa_qtd_agricul_familiar_sexo_feminino_i": "sum",
        "paa_vlr_pago_exec_compra_doacao_simul_d": "sum",
        "paa_vlr_pago_exec_incentivo_leite_d": "sum"
    }).reset_index()
    
    # --- NIVEL 3: RN TERRITORIAL ---
    def get_territory(ibge):
        ibge_int = int(str(ibge)[:6])
        for t, codes in TERRITORIOS_RN.items():
            if ibge_int in codes: return t
        return "N/M"
    
    df_rn['Território'] = df_rn['codigo_ibge'].apply(get_territory)
    rn_territorial = df_rn[df_rn['Território'] != "N/M"].groupby(['Território', 'Ano']).agg({
        "recur_pagos_agricul_paa_f": "sum",
        "agricultores_fornec_paa_i": "sum",
        "paa_qtd_agricul_familiar_sexo_feminino_i": "sum"
    }).reset_index()
    
    # --- NIVEL 4: RN MUNICIPAL ---
    # Busca nomes IBGE
    mun_resp = requests.get("https://servicodados.ibge.gov.br/api/v1/localidades/estados/24/municipios").json()
    mapping = {str(m['id'])[:6]: m['nome'] for m in mun_resp}
    df_rn['Município'] = df_rn['codigo_ibge'].astype(str).str[:6].map(mapping)
    rn_municipal = df_rn.groupby(['Município', 'Território', 'Ano']).agg({
        "recur_pagos_agricul_paa_f": "sum",
        "agricultores_fornec_paa_i": "sum",
        "paa_qtd_agricul_familiar_sexo_feminino_i": "sum",
        "paa_vlr_pago_exec_incentivo_leite_d": "sum"
    }).reset_index()

    # Salvando Tudo
    with pd.ExcelWriter(OUTPUT_MASTER, engine='openpyxl') as writer:
        ne_comparativo.to_excel(writer, sheet_name="1. Nordeste Comparativo", index=False)
        format_excel(writer, "1. Nordeste Comparativo", ne_comparativo)
        
        rn_geral.to_excel(writer, sheet_name="2. RN Geral", index=False)
        format_excel(writer, "2. RN Geral", rn_geral)
        
        rn_territorial.to_excel(writer, sheet_name="3. RN Territorial", index=False)
        format_excel(writer, "3. RN Territorial", rn_territorial)
        
        rn_municipal.to_excel(writer, sheet_name="4. RN Municipal", index=False)
        format_excel(writer, "4. RN Municipal", rn_municipal)

    print(f"✨ Missão Cumprida! Relatório Master Final: {OUTPUT_MASTER}")

if __name__ == "__main__":
    generate()
