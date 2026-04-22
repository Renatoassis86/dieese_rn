"""
Extrator da Base 1 - PAA Execução Financeira via Termo de Adesão
Fonte: API Portal da Transparência (CGU)
Variáveis cobertas: Valor Pactuado, Valor Pago, Valor Não Executado,
                   % Execução, Status, Número do Convênio, Vigência,
                   Município, UF, Tipo Adesão
"""
import os
import requests
import urllib3
import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

urllib3.disable_warnings()
load_dotenv()

API_KEY = os.getenv('PORTAL_TRANSPARENCIA_API_KEY', 'e5874effcecce2dce606fedcf6c16471')
HEADERS = {'chave-api-dados': API_KEY, 'Accept': 'application/json'}
BASE = 'https://api.portaldatransparencia.gov.br/api-de-dados'
OUTPUT_FILE = 'outputs/paa_reports/PAA_Base1_Pactuacao_TermoAdesao_RN.xlsx'
os.makedirs('outputs/paa_reports', exist_ok=True)

# Paleta DIEESE
COLOR_NAVY = "0D1B2A"
COLOR_WHITE = "FFFFFF"
COLOR_GRAY_LT = "F2F2F2"
COLOR_GREEN = "2D6A4F"


def fetch_convenios_rn_mds(year: int) -> list:
    """Busca convênios do MDS para RN em um determinado ano."""
    all_records = []
    page = 1
    while True:
        try:
            r = requests.get(
                f'{BASE}/convenios',
                params={
                    'uf': 'RN',
                    'codigoOrgao': '55000',
                    'dataVigenciaInicial': f'01/01/{year}',
                    'dataVigenciaFinal': f'31/12/{year}',
                    'pagina': page
                },
                headers=HEADERS, verify=False, timeout=30
            )
            if r.status_code != 200 or not r.json():
                break
            data = r.json()
            all_records.extend(data)
            if len(data) < 20:  # Página incompleta = última página
                break
            page += 1
        except Exception as e:
            print(f'  Erro na página {page}/{year}: {e}')
            break
    return all_records


def is_paa_related(objeto: str, subfuncao: str = '') -> bool:
    """Verifica se o objeto do convênio é relacionado ao PAA ou alimentação."""
    texto = (objeto + ' ' + subfuncao).upper()
    keywords = [
        'AQUISI', 'ALIMENTO', 'PAA', 'PROGRAMA DE AQUISICAO',
        'DOACAO SIMULTANEA', 'LEITE', 'AGRICULTURA FAMILIAR',
        'FORNECIMENTO', 'DISTRIBUICAO DE ALIMENTOS', 'SEGURANCA ALIMENTAR',
        'SEMENTES', 'ESTOQUE', 'PRODUCAO RURAL', 'FAMILIAR'
    ]
    return any(k in texto for k in keywords)


def format_sheet(writer, sheet_name, df):
    ws = writer.sheets[sheet_name]
    # Cabeçalho
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(name='Montserrat', bold=True, color=COLOR_WHITE, size=10)
        cell.fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Dados
    for row_num in range(2, len(df) + 2):
        bg = COLOR_GRAY_LT if row_num % 2 == 0 else COLOR_WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            col_name = df.columns[col_num - 1].lower()
            if 'r$' in col_name or 'valor' in col_name:
                cell.number_format = 'R$ #,##0.00'
            elif '%' in col_name:
                cell.number_format = '0.0%'

    # Largura de colunas
    for i, col in enumerate(df.columns, 1):
        col_data = df[col].tolist()
        lengths = [len(str(x)) for x in col_data if x is not None]
        max_len = max(lengths) if lengths else 0
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len, len(col)) + 4, 55)

    # Linha de cabeçalho congelada
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 40


def generate_paa_base1():
    print('=' * 60)
    print('BASE 1 — PAA Execução Financeira via Termo de Adesão')
    print('Fonte: API Portal da Transparência (CGU)')
    print('=' * 60)

    all_convenios = []
    years = range(2021, 2027)  # Documento cita 2024-2026, mas ampliamos para capturar mais

    for year in years:
        print(f'Buscando convênios MDS/RN em {year}...')
        records = fetch_convenios_rn_mds(year)
        print(f'  Total bruto: {len(records)} convênios')
        
        paa_records = []
        for r in records:
            obj = r.get('dimConvenio', {}).get('objeto', '') or ''
            sub = r.get('subfuncao', {}).get('descricaoSubfuncap', '') or r.get('subfuncao', {}).get('descricaoSubfuncao', '') or ''
            if is_paa_related(obj, sub):
                paa_records.append(r)
        print(f'  Relacionados ao PAA/Alimentação: {len(paa_records)}')
        all_convenios.extend(paa_records)

    if not all_convenios:
        print('\n⚠️  Nenhum convênio PAA específico encontrado via filtro automático.')
        print('   Incluindo TODOS os convênios MDS/RN para análise completa...')
        for year in range(2021, 2027):
            records = fetch_convenios_rn_mds(year)
            all_convenios.extend(records)

    # Deduplica por ID
    seen = set()
    unique = []
    for c in all_convenios:
        if c['id'] not in seen:
            seen.add(c['id'])
            unique.append(c)

    print(f'\nTotal de convênios únicos coletados: {len(unique)}')

    # Montar DataFrame estruturado com as variáveis do Documento
    rows = []
    for c in unique:
        convenio = c.get('dimConvenio', {}) or {}
        municipio = c.get('municipioConvenente', {}) or {}
        ibge = municipio.get('codigoIBGE', '') or ''
        mun_nome = municipio.get('nomeIBGE', '') or municipio.get('nome', '')
        
        convenente = c.get('convenente', {}) or {}
        if isinstance(convenente, dict):
            convenente_nome = convenente.get('nome', '')
        else:
            convenente_nome = str(convenente)
        
        orgao_dict = c.get('orgao', {}) or {}
        orgao_nome = orgao_dict.get('nome', '') if isinstance(orgao_dict, dict) else str(orgao_dict)
        
        subfuncao_dict = c.get('subfuncao', {}) or {}
        subfuncao_nome = ''
        if isinstance(subfuncao_dict, dict):
            subfuncao_nome = subfuncao_dict.get('descricaoSubfuncap', '') or subfuncao_dict.get('descricao', '')
        
        tipo_inst = c.get('tipoInstrumento', {}) or {}
        tipo_inst_nome = tipo_inst.get('descricao', '') if isinstance(tipo_inst, dict) else str(tipo_inst)
        valor_total = c.get('valor', 0) or 0
        valor_liberado = c.get('valorLiberado', 0) or 0
        valor_contrapartida = c.get('valorContrapartida', 0) or 0
        valor_nao_exec = max(valor_total - valor_liberado, 0)
        perc_exec = (valor_liberado / valor_total * 100) if valor_total > 0 else 0

        row = {
            'Código IBGE': ibge,
            'UF': 'RN',
            'Município': mun_nome,
            'Convenente (Executor)': convenente_nome,
            'Tipo Instrumento': tipo_inst_nome,
            'Número Convênio': convenio.get('numero', ''),
            'Objeto / Descrição': convenio.get('objeto', ''),
            'Subfunção': subfuncao_nome,
            'Órgão Concedente': orgao_nome,
            'Situação': c.get('situacao', ''),
            'Data Início Vigência': c.get('dataInicioVigencia', ''),
            'Data Fim Vigência': c.get('dataFinalVigencia', ''),
            'Data Publicação': c.get('dataPublicacao', ''),
            'Data Última Liberação': c.get('dataUltimaLiberacao', ''),
            'Valor Total (R$)': valor_total,
            'Valor Liberado/Pago (R$)': valor_liberado,
            'Valor Contrapartida (R$)': valor_contrapartida,
            'Valor Não Executado (R$)': valor_nao_exec,
            '% Execução': perc_exec,
            'Número Processo': c.get('numeroProcesso', ''),
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # Organizar por município e data
    df = df.sort_values(['Município', 'Data Início Vigência'], ascending=[True, False])

    print(f'\nSalvando em: {OUTPUT_FILE}')
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='1. Pactuação Termo Adesão', index=False)
        format_sheet(writer, '1. Pactuação Termo Adesão', df)

        # Resumo por município
        df_resumo = df.groupby('Município').agg(
            Total_Convenios=('Número Convênio', 'count'),
            Valor_Total_R=('Valor Total (R$)', 'sum'),
            Valor_Pago_R=('Valor Liberado/Pago (R$)', 'sum'),
            Valor_Nao_Exec_R=('Valor Não Executado (R$)', 'sum'),
        ).reset_index()
        df_resumo['% Execução'] = (df_resumo['Valor_Pago_R'] / df_resumo['Valor_Total_R'] * 100).fillna(0)
        df_resumo.columns = ['Município', 'Qtd Convênios', 'Valor Total (R$)', 'Valor Pago (R$)', 'Valor Não Exec. (R$)', '% Execução']
        df_resumo.to_excel(writer, sheet_name='Resumo por Município', index=False)
        format_sheet(writer, 'Resumo por Município', df_resumo)

    print(f'✅ Gerado com sucesso! {len(df)} registros em {len(df["Município"].unique())} municípios.')
    return OUTPUT_FILE


if __name__ == '__main__':
    generate_paa_base1()
