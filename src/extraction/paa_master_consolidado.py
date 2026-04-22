import pandas as pd
import requests
import io
import os
import time
import urllib3
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configurações
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
OUTPUT_FILE = "outputs/paa_reports/PAA_RN_MASTER_CONSOLIDADO.xlsx"
os.makedirs("outputs/paa_reports", exist_ok=True)

# API Keys e Endpoints
API_KEY_PORTAL = 'e5874effcecce2dce606fedcf6c16471'
HEADERS_PORTAL = {
    'chave-api-dados': API_KEY_PORTAL, 
    'Accept': 'application/json',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}
BASE_PORTAL = 'https://api.portaldatransparencia.gov.br/api-de-dados'
MDS_API_URL = "https://aplicacoes.mds.gov.br/sagi/servicos/misocial/"

# Cores DIEESE
COLOR_NAVY = "0D1B2A"
COLOR_WHITE = "FFFFFF"
COLOR_GRAY_LT = "F2F2F2"

def get_municipality_mapping():
    url = "https://servicodados.ibge.gov.br/api/v1/localidades/estados/24/municipios"
    try:
        resp = requests.get(url, timeout=30)
        return {str(m['id'])[:6]: m['nome'] for m in resp.json()}
    except:
        return {}

def format_excel_sheet(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    header_font = Font(name='Calibri', bold=True, color=COLOR_WHITE)
    header_fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, len(df) + 2):
        fill = PatternFill(start_color=COLOR_GRAY_LT, end_color=COLOR_GRAY_LT, fill_type="solid") if row % 2 == 0 else None
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            if fill: cell.fill = fill
            col_name = df.columns[col-1].lower()
            if "r$" in col_name or "valor" in col_name or "pago" in col_name:
                cell.number_format = 'R$ #,##0.00'
            elif "kg" in col_name or "litros" in col_name or "quantidade" in col_name:
                cell.number_format = '#,##0.00'

    for i, col in enumerate(df.columns):
        ws_col = worksheet.column_dimensions[get_column_letter(i + 1)]
        ws_col.width = 25

def fetch_portal_convenios():
    """Busca Base 1 via Portal da Transparência RN."""
    all_paa = []
    print("Extraindo Base 1 (Portal da Transparência)...")
    for year in range(2021, 2027):
        page = 1
        while True:
            try:
                r = requests.get(
                    f'{BASE_PORTAL}/convenios',
                    params={'uf': 'RN', 'codigoOrgao': '55000', 'dataVigenciaInicial': f'01/01/{year}', 'dataVigenciaFinal': f'31/12/{year}', 'pagina': page},
                    headers=HEADERS_PORTAL, verify=False, timeout=30
                )
                if r.status_code != 200 or not r.json(): break
                data = r.json()
                for item in data:
                    obj = (item.get('dimConvenio', {}).get('objeto', '') or '').upper()
                    if any(k in obj for k in ['PAA', 'AQUISICAO DE ALIMENTO', 'LEITE', 'AGRICULTURA FAMILIAR']):
                        all_paa.append(item)
                if len(data) < 15: break
                page += 1
            except: break
    
    rows = []
    for c in all_paa:
        municipio = c.get('municipioConvenente', {}) or {}
        val_total = c.get('valor', 0) or 0
        val_lib = c.get('valorLiberado', 0) or 0
        rows.append({
            'Município': municipio.get('nomeIBGE', ''),
            'Objeto': c.get('dimConvenio', {}).get('objeto', ''),
            'Valor Total (R$)': val_total,
            'Valor Pago (R$)': val_lib,
            'Vigência Início': c.get('dataInicioVigencia', ''),
            'Vigência Fim': c.get('dataFinalVigencia', ''),
            'Situação': c.get('situacao', '')
        })
    return pd.DataFrame(rows).drop_duplicates()

def generate_master_report():
    print("="*60)
    print("GERADOR MASTER PAA RN - CONSOLIDAÇÃO FINAL (PAA API + PORTAL)")
    print("="*60)
    
    mapping = get_municipality_mapping()
    
    # 1. Base 1 (Portal)
    df_base1 = fetch_portal_convenios()
    
    # 2. Bases 2-11 (MiSocial)
    fields = [
        "codigo_ibge", "anomes_s", "paa_indicador_adesao_municipio_i",
        "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f",
        "paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i",
        "paa_qtd_agricul_familiar_exec_compra_doacao_simul_i", "paa_vlr_pago_exec_compra_doacao_simul_d", "paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d",
        "paa_qtd_agricul_familiar_exec_incentivo_leite_i", "paa_vlr_pago_exec_incentivo_leite_d", "paa_qtd_alim_adquiridos_exec_incentivo_leite_i",
        "paa_qtd_agricul_familiar_modal_compra_doacao_simul_i", "paa_qtd_agricul_familiar_modal_incentivo_leite_i",
        "paa_qtd_agricul_familiar_modal_compra_direta_i", "paa_qtd_agricul_familiar_modal_formacao_estoque_i", "paa_qtd_agricul_familiar_modal_aquisicao_sementes_i"
    ]
    
    print("Extraindo Bases 2-11 (API MiSocial)...")
    url_misocial = f"{MDS_API_URL}?q=codigo_ibge:24*&rows=1000000&wt=csv&fl={','.join(fields)}"
    resp = requests.get(url_misocial, verify=False, timeout=120)
    df_full = pd.read_csv(io.StringIO(resp.text))
    df_full['Ano'] = df_full['anomes_s'].astype(str).str[:4]
    df_full['Município'] = df_full['codigo_ibge'].astype(str).str[:6].map(mapping)
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Sheet Base 1
        if not df_base1.empty:
            df_base1.to_excel(writer, sheet_name="1. Pactuação (Transparência)", index=False)
            format_excel_sheet(writer, "1. Pactuação (Transparência)", df_base1)
        
        # Mapping Bases
        bases = {
            "2. Adesão Municipal": (["paa_indicador_adesao_municipio_i"], {"paa_indicador_adesao_municipio_i": "max"}),
            "3. Execução Geral": (["agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f"], {"agricultores_fornec_paa_i": "sum", "recur_pagos_agricul_paa_f": "sum"}),
            "4. Perfil Sexo": (["paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i"], {"paa_qtd_agricul_familiar_sexo_masculino_i": "sum", "paa_qtd_agricul_familiar_sexo_feminino_i": "sum"}),
            "5. Doação Simultânea": (["paa_qtd_agricul_familiar_exec_compra_doacao_simul_i", "paa_vlr_pago_exec_compra_doacao_simul_d"], {"paa_qtd_agricul_familiar_exec_compra_doacao_simul_i": "sum", "paa_vlr_pago_exec_compra_doacao_simul_d": "sum"}),
            "6. Modalidade Leite": (["paa_qtd_agricul_familiar_exec_incentivo_leite_i", "paa_vlr_pago_exec_incentivo_leite_d"], {"paa_qtd_agricul_familiar_exec_incentivo_leite_i": "sum", "paa_vlr_pago_exec_incentivo_leite_d": "sum"}),
            "7. Agricultores Modalidades": (["paa_qtd_agricul_familiar_modal_compra_doacao_simul_i", "paa_qtd_agricul_familiar_modal_incentivo_leite_i", "paa_qtd_agricul_familiar_modal_compra_direta_i", "paa_qtd_agricul_familiar_modal_formacao_estoque_i", "paa_qtd_agricul_familiar_modal_aquisicao_sementes_i"], {"paa_qtd_agricul_familiar_modal_compra_doacao_simul_i": "sum", "paa_qtd_agricul_familiar_modal_incentivo_leite_i": "sum", "paa_qtd_agricul_familiar_modal_compra_direta_i": "sum", "paa_qtd_agricul_familiar_modal_formacao_estoque_i": "sum", "paa_qtd_agricul_familiar_modal_aquisicao_sementes_i": "sum"})
        }
        
        for name, (cols, agg) in bases.items():
            print(f"Processando {name}...")
            mask = df_full[cols].sum(axis=1) > 0
            df_slice = df_full[mask].copy()
            if not df_slice.empty:
                df_sheet = df_slice.groupby(['Município', 'Ano']).agg(agg).reset_index()
                df_sheet.columns = [c.replace('paa_', '').replace('_i', '').replace('_f', '').replace('_d', '').replace('_', ' ').title() for c in df_sheet.columns]
                df_sheet.to_excel(writer, sheet_name=name, index=False)
                format_excel_sheet(writer, name, df_sheet)

    print(f"\n✅ Relatório Master Gerado: {OUTPUT_FILE}")
    print(f"Municípios cobertos (API PAA): {df_full['Município'].nunique()}")

if __name__ == "__main__":
    generate_master_report()
