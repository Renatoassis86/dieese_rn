import os, io, json, requests, urllib3, pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()

# ─── Configurações ────────────────────────────────────────────────────────────
MDS_API   = "https://aplicacoes.mds.gov.br/sagi/servicos/misocial/"
OUTPUT    = "outputs/paa_reports/PAA_ESTUDO_COMPARATIVO_TERRITORIAL.xlsx"
IBGE_JSON = "data/bronze/ibge/municipios_rn.json"
os.makedirs("outputs/paa_reports", exist_ok=True)

# ─── Mapeamento de Estados Nordeste ───────────────────────────────────────────
NE_STATES = {
    "21": "Maranhão",
    "22": "Piauí",
    "23": "Ceará",
    "24": "Rio Grande do Norte",
    "25": "Paraíba",
    "26": "Pernambuco",
    "27": "Alagoas",
    "28": "Sergipe",
    "29": "Bahia"
}

# ─── 10 Territórios do RN ─────────────────────────────────────────────────────
TERRITORIOS = {
    "Seridó": [240350, 240400, 240450, 240460, 240600, 240650, 240660, 240840, 240900, 241010, 241020, 241030, 240980, 241170, 241350, 241420, 241570, 241575, 241580, 241585, 241590, 241600, 241602, 241680, 241690],
    "Alto Oeste": [240050, 240100, 240110, 240120, 240310, 240370, 240560, 240630, 240640, 241040, 241050, 241090, 241100, 241130, 241140, 241145, 241390, 241400, 241405, 241410, 241470, 241480, 241485, 241560, 241565, 241610, 241650, 241660, 241760, 241770],
    "Agreste Litoral Sul": [240070, 240180, 240190, 240330, 240540, 240700, 241060, 241110, 241115, 241120, 241160, 241165, 241230, 241240, 241430, 241440, 241520, 241525, 241540, 241545, 241670, 241740, 241750, 241755],
    "Mato Grande": [240200, 240290, 240300, 241080, 241085, 241088, 241200, 241380, 241385, 241388, 241395, 241490, 241530, 241535, 241640, 241645, 241735, 241745, 241765],
    "Assú-Mossoró": [240060, 240065, 240067, 240210, 240340, 240710, 240950, 240960, 241150, 241415, 241420, 241500, 241505, 241700],
    "Potengi": [240220, 240230, 240970, 241075, 241455, 241460, 241510, 241515, 241518, 241519, 241521],
    "Trairi": [240240, 240280, 240285, 241070, 241072, 241095, 241155, 241380, 241446, 241550, 241555, 241556, 241557, 241558, 241720],
    "Sertão Central Cabugi e Litoral Norte": [240080, 240090, 240320, 240580, 240720, 240730, 241105, 241210, 241445, 241450],
    "Sertão do Apodi": [240130, 240140, 240360, 240590, 240740, 240960, 241065, 241220, 241340, 241375, 241376, 241460, 241465, 241550, 241710, 241770, 241780],
    "Terras Potiguares": [240570, 241180, 241250, 241430, 241555]
}

# ─── Cores DIEESE ────────────────────────────────────────────────────────────
NAVY   = "0D1B2A"
WHITE  = "FFFFFF"
ZEBRA  = "EFF3F8"

def format_sheet(ws, df):
    """Formata uma sheet com padrão DIEESE."""
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
        cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_num in range(2, len(df) + 2):
        bg = ZEBRA if row_num % 2 == 0 else WHITE
        fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            col_name = df.columns[col_num - 1]
            if 'R$' in col_name or 'Valor' in col_name or 'Recurso' in col_name:
                cell.number_format = 'R$ #,##0.00'
            elif 'Ticket' in col_name:
                cell.number_format = 'R$ #,##0.00'
            elif '%' in col_name:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0'

    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 25

def generate_comparativo():
    print("=" * 65)
    print("PAA RN — ESTUDO TERRITORIAL E COMPARATIVO NORDESTE")
    print("=" * 65)

    # 1. Mapeamento Territorial RN
    with open(IBGE_JSON, encoding='utf-8') as f:
        munis_rn = json.load(f)
    ibge_nome = {str(m['id'])[:6]: m['nome'] for m in munis_rn}
    ibge_to_territorio = {}
    for terr, cods in TERRITORIOS.items():
        for cod in cods:
            ibge_to_territorio[str(cod)] = terr

    # 2. Extração Nordeste (Execução Geral - Base 3)
    print("\n[1/4] Extraindo dados do Nordeste (API MiSocial)...")
    fields_ne = ["codigo_ibge", "anomes_s", "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f"]
    # Filtro: prefixos IBGE do Nordeste (21 a 29)
    # A API permite q=codigo_ibge:2* (Pega todo o NE + SE + CO). Vamos pegar 2* e filtrar localmente.
    url_ne = f"{MDS_API}?q=codigo_ibge:2*&rows=1000000&wt=csv&fl={','.join(fields_ne)}"
    resp = requests.get(url_ne, verify=False, timeout=180)
    df_raw_ne = pd.read_csv(io.StringIO(resp.text))
    df_raw_ne['UF_Code'] = df_raw_ne['codigo_ibge'].astype(str).str[:2]
    df_raw_ne = df_raw_ne[df_raw_ne['UF_Code'].isin(NE_STATES.keys())].copy()
    df_raw_ne['UF'] = df_raw_ne['UF_Code'].map(NE_STATES)
    df_raw_ne['Ano'] = df_raw_ne['anomes_s'].astype(str).str[:4]
    
    # Agregação Nordeste
    ne_comparativo = df_raw_ne.groupby(['UF', 'Ano']).agg(
        Qtd_Agricultores=('agricultores_fornec_paa_i', 'sum'),
        Valor_Pago=('recur_pagos_agricul_paa_f', 'sum')
    ).reset_index()
    ne_comparativo = ne_comparativo[ne_comparativo['Valor_Pago'] > 0].copy()
    ne_comparativo['Ticket Médio (R$)'] = ne_comparativo['Valor_Pago'] / ne_comparativo['Qtd_Agricultores']
    ne_comparativo.rename(columns={'Valor_Pago': 'Valor Pago (R$)', 'Qtd_Agricultores': 'Agricultores Fornecedores'}, inplace=True)

    # 3. Extração Detalhada RN (Todas as Bases)
    print("\n[2/4] Extraindo dados detalhados do RN (Todas as Bases)...")
    fields_rn = [
        "codigo_ibge", "anomes_s",
        "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f",
        "paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i",
        "paa_qtd_agricul_familiar_exec_compra_doacao_simul_i", "paa_vlr_pago_exec_compra_doacao_simul_d",
        "paa_qtd_agricul_familiar_exec_incentivo_leite_i", "paa_vlr_pago_exec_incentivo_leite_d"
    ]
    url_rn = f"{MDS_API}?q=codigo_ibge:24*&rows=500000&wt=csv&fl={','.join(fields_rn)}"
    resp_rn = requests.get(url_rn, verify=False, timeout=120)
    df_rn = pd.read_csv(io.StringIO(resp_rn.text))
    df_rn['ibge6'] = df_rn['codigo_ibge'].astype(str).str[:6]
    df_rn['Ano'] = df_rn['anomes_s'].astype(str).str[:4]
    df_rn['Território'] = df_rn['ibge6'].map(ibge_to_territorio).fillna('Interior/Outros')
    
    # Agregação RN Geral
    rn_geral = df_rn.groupby('Ano').agg(
        Agricultores=('agricultores_fornec_paa_i', 'sum'),
        Valor_Total=('recur_pagos_agricul_paa_f', 'sum'),
        Mulheres=('paa_qtd_agricul_familiar_sexo_feminino_i', 'sum'),
        Recurso_Leite=('paa_vlr_pago_exec_incentivo_leite_d', 'sum')
    ).reset_index()
    rn_geral = rn_geral[rn_geral['Valor_Total'] > 0].copy()
    rn_geral['% Mulheres'] = rn_geral['Mulheres'] / rn_geral['Agricultores']
    rn_geral.rename(columns={'Valor_Total': 'Valor Total (R$)', 'Recurso_Leite': 'Recurso Leite (R$)'}, inplace=True)

    # Agregação RN Territorial
    print("\n[3/4] Agregando dados pelos 10 territórios do RN...")
    rn_territorial = df_rn.groupby(['Território', 'Ano']).agg(
        Agricultores=('agricultores_fornec_paa_i', 'sum'),
        Valor_Total=('recur_pagos_agricul_paa_f', 'sum'),
        Mulheres=('paa_qtd_agricul_familiar_sexo_feminino_i', 'sum'),
        Quantidade_Municipios=('ibge6', 'nunique')
    ).reset_index()
    rn_territorial = rn_territorial[rn_territorial['Valor_Total'] > 0].copy()
    rn_territorial['Ticket Médio (R$)'] = rn_territorial['Valor_Total'] / rn_territorial['Agricultores']
    rn_territorial['% Mulheres'] = rn_territorial['Mulheres'] / rn_territorial['Agricultores']
    rn_territorial.rename(columns={'Valor_Total': 'Valor Total (R$)'}, inplace=True)

    # 4. Gerar Excel
    print(f"\n[4/4] Gerando estudo em {OUTPUT}...")
    with pd.ExcelWriter(OUTPUT, engine='openpyxl') as writer:
        # 1. Comparativo Nordeste
        ne_comparativo.to_excel(writer, sheet_name='1. Nordeste Comparativo', index=False)
        format_sheet(writer.sheets['1. Nordeste Comparativo'], ne_comparativo)
        
        # 2. RN Geral
        rn_geral.to_excel(writer, sheet_name='2. RN Geral', index=False)
        format_sheet(writer.sheets['2. RN Geral'], rn_geral)
        
        # 3. RN Territorial
        rn_territorial.to_excel(writer, sheet_name='3. RN Territorial', index=False)
        format_sheet(writer.sheets['3. RN Territorial'], rn_territorial)

    print(f"\n✅ Estudo gerado com sucesso!")
    print(f"   Arquivo: {OUTPUT}")

if __name__ == "__main__":
    generate_comparativo()
