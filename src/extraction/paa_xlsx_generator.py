import pandas as pd
import requests
import io
import os
import time
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configurações
OUTPUT_FILE = "outputs/paa_reports/PAA_VALIDACAO_COMPLETA_11_BASES.xlsx"
os.makedirs("outputs/paa_reports", exist_ok=True)

# Cores DIEESE
COLOR_NAVY = "0D1B2A"
COLOR_WHITE = "FFFFFF"
COLOR_GRAY_LT = "F2F2F2"

def get_municipality_mapping():
    url = "https://servicodados.ibge.gov.br/api/v1/localidades/estados/24/municipios"
    try:
        resp = requests.get(url)
        return {str(m['id'])[:6]: m['nome'] for m in resp.json()}
    except:
        return {}

def format_excel_sheet(writer, sheet_name, df, title):
    worksheet = writer.sheets[sheet_name]
    header_font = Font(name='Montserrat', bold=True, color=COLOR_WHITE)
    header_fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    
    # Header format
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for i, col in enumerate(df.columns):
        col_data = df[col].tolist()
        lengths = [len(str(x)) for x in col_data if x is not None]
        max_data_len = max(lengths) if lengths else 0
        column_len = max(max_data_len, len(str(col))) + 4
        worksheet.column_dimensions[get_column_letter(i + 1)].width = min(column_len, 60)

    # Zebra stripes and number formatting
    for row in range(2, len(df) + 2):
        fill = PatternFill(start_color=COLOR_GRAY_LT, end_color=COLOR_GRAY_LT, fill_type="solid") if row % 2 == 0 else None
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            if fill: cell.fill = fill
            
            col_name = df.columns[col-1].lower()
            if "r$" in col_name or "valor" in col_name or "pago" in col_name:
                cell.number_format = '"R$ "#,##0.00'
            elif "kg" in col_name or "litros" in col_name or "quantidade" in col_name:
                cell.number_format = '"#,##0.00'

def validate_and_generate():
    print("Iniciando Validação e Geração das 11 Bases PAA...")
    
    # 1. Carregar Mapeamento
    mapping = get_municipality_mapping()
    
    # 2. Extração de Dados (MiSocial contém as bases 2 a 11)
    fields = [
        "codigo_ibge", "anomes_s",
        "paa_indicador_adesao_municipio_i",
        "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f",
        "paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i", "paa_qtd_agricul_familiar_sem_info_sexo_i",
        "paa_qtd_agricul_familiar_exec_compra_doacao_simul_i", "paa_vlr_pago_exec_compra_doacao_simul_d", "paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d",
        "paa_qtd_agricul_familiar_exec_incentivo_leite_i", "paa_vlr_pago_exec_incentivo_leite_d", "paa_qtd_alim_adquiridos_exec_incentivo_leite_i",
        "paa_qtd_agricul_familiar_modal_compra_doacao_simul_i",
        "paa_qtd_agricul_familiar_modal_incentivo_leite_i",
        "paa_qtd_agricul_familiar_modal_compra_direta_i",
        "paa_qtd_agricul_familiar_modal_formacao_estoque_i",
        "paa_qtd_agricul_familiar_modal_aquisicao_sementes_i"
    ]
    
    print("Buscando dados na API MiSocial...")
    url = f"https://aplicacoes.mds.gov.br/sagi/servicos/misocial/?q=codigo_ibge:24*&rows=1000000&wt=csv&fl={','.join(fields)}"
    
    try:
        resp = requests.get(url, verify=False, timeout=120)
        df_full = pd.read_csv(io.StringIO(resp.text))
        df_full['Ano'] = df_full['anomes_s'].astype(str).str[:4]
        df_full['Município'] = df_full['codigo_ibge'].astype(str).map(mapping)
    except Exception as e:
        print(f"Erro crítico na extração MiSocial: {e}")
        return

    # 3. Definição das Bases (Documento vs API)
    definitions = {
        "2. Adesão Municipal": {
            "fields": ["paa_indicador_adesao_municipio_i"],
            "period_doc": "2022-2026",
            "agg": {"paa_indicador_adesao_municipio_i": "max"}
        },
        "3. Execução Geral": {
            "fields": ["agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f"],
            "period_doc": "2011-2025",
            "agg": {"agricultores_fornec_paa_i": "sum", "recur_pagos_agricul_paa_f": "sum"}
        },
        "4. Perfil Sexo": {
            "fields": ["paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i", "paa_qtd_agricul_familiar_sem_info_sexo_i"],
            "period_doc": "2014-2025",
            "agg": {"paa_qtd_agricul_familiar_sexo_masculino_i": "sum", "paa_qtd_agricul_familiar_sexo_feminino_i": "sum", "paa_qtd_agricul_familiar_sem_info_sexo_i": "sum"}
        },
        "5. Termo Adesão (Exec)": {
            "fields": ["paa_qtd_agricul_familiar_exec_compra_doacao_simul_i", "paa_vlr_pago_exec_compra_doacao_simul_d", "paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d"],
            "period_doc": "2021-2025",
            "agg": {"paa_qtd_agricul_familiar_exec_compra_doacao_simul_i": "sum", "paa_vlr_pago_exec_compra_doacao_simul_d": "sum", "paa_qtd_alim_adquiridos_exec_compra_doacao_simul_d": "sum"}
        },
        "6. Modalidade Leite": {
            "fields": ["paa_qtd_agricul_familiar_exec_incentivo_leite_i", "paa_vlr_pago_exec_incentivo_leite_d", "paa_qtd_alim_adquiridos_exec_incentivo_leite_i"],
            "period_doc": "2021-2025",
            "agg": {"paa_qtd_agricul_familiar_exec_incentivo_leite_i": "sum", "paa_vlr_pago_exec_incentivo_leite_d": "sum", "paa_qtd_alim_adquiridos_exec_incentivo_leite_i": "sum"}
        },
        "7. Fornec. Doação Simut.": {
            "fields": ["paa_qtd_agricul_familiar_modal_compra_doacao_simul_i"],
            "period_doc": "2011-2025",
            "agg": {"paa_qtd_agricul_familiar_modal_compra_doacao_simul_i": "sum"}
        },
        "8. Fornec. Leite": {
            "fields": ["paa_qtd_agricul_familiar_modal_incentivo_leite_i"],
            "period_doc": "2011-2025",
            "agg": {"paa_qtd_agricul_familiar_modal_incentivo_leite_i": "sum"}
        },
        "9. Compra Direta": {
            "fields": ["paa_qtd_agricul_familiar_modal_compra_direta_i"],
            "period_doc": "2011-2017",
            "agg": {"paa_qtd_agricul_familiar_modal_compra_direta_i": "sum"}
        },
        "10. Apoio Estoque": {
            "fields": ["paa_qtd_agricul_familiar_modal_formacao_estoque_i"],
            "period_doc": "2011-2019",
            "agg": {"paa_qtd_agricul_familiar_modal_formacao_estoque_i": "sum"}
        },
        "11. Sementes": {
            "fields": ["paa_qtd_agricul_familiar_modal_aquisicao_sementes_i"],
            "period_doc": "2015-2020",
            "agg": {"paa_qtd_agricul_familiar_modal_aquisicao_sementes_i": "sum"}
        }
    }

    results = []
    summary_data = []

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        for name, cfg in definitions.items():
            print(f"Processando {name}...")
            
            # Filtro por disponibilidade real na API
            mask = df_full[cfg['fields']].sum(axis=1) > 0
            df_slice = df_full[mask].copy()
            
            if df_slice.empty:
                print(f"⚠️ Sem dados para {name}")
                summary_data.append({"Métrica": name, "Período Doc": cfg['period_doc'], "Período API": "SEM DADOS", "Status": "FALHA"})
                continue
            
            min_api = df_slice['Ano'].min()
            max_api = df_slice['Ano'].max()
            summary_data.append({"Métrica": name, "Período Doc": cfg['period_doc'], "Período API": f"{min_api}-{max_api}", "Status": "OK"})

            # Agregação
            df_sheet = df_slice.groupby(['Município', 'Ano']).agg(cfg['agg']).reset_index()
            
            # Limpeza visual de nomes de colunas
            df_sheet.columns = [c.replace('paa_', '').replace('_i', '').replace('_f', '').replace('_d', '').replace('_', ' ').title() for c in df_sheet.columns]
            
            df_sheet.to_excel(writer, sheet_name=name[:30], index=False)
            format_excel_sheet(writer, name[:30], df_sheet, name)

        # Base 1 - Especial (Termo Adesão Pactuação)
        # Como não temos os campos pactuados no MiSocial, criamos uma aba de aviso/placeholder com o que temos
        print("Processando 1. Pactuação (Termo Adesão)...")
        summary_data.insert(0, {"Métrica": "1. Pactuação Termo Adesão", "Período Doc": "2024-2026", "Período API": "Não em MiSocial", "Status": "Requer Base Específica"})
        
        # Aba de Resumo
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='0. Resumo Validação', index=False)
        format_excel_sheet(writer, '0. Resumo Validação', df_summary, "Resumo de Validação Documento vs API")

    print(f"Processo Concluído. Arquivo: {OUTPUT_FILE}")

if __name__ == "__main__":
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    validate_and_generate()
