import pandas as pd
import requests
import io
import os
import urllib3
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Configurações
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
OUTPUT_FILE = "outputs/paa_reports/PAA_DIAGNOSTICO_ESTUDO_RN.xlsx"
MDS_API_URL = "https://aplicacoes.mds.gov.br/sagi/servicos/misocial/"
os.makedirs("outputs/paa_reports", exist_ok=True)

# 10 Territórios do RN (Mapeamento IBGE 6 dígitos)
TERRITORIOS = {
    "Seridó": [240350, 240400, 240450, 240460, 240600, 240650, 240660, 240840, 240900, 241010, 241020, 241030, 240980, 241170, 241350, 241420, 241570, 241575, 241580, 241585, 241590, 241600, 241602, 241680, 241690],
    "Alto Oeste": [240050, 240100, 240110, 240120, 240310, 240370, 240560, 240630, 240640, 241040, 241050, 241090, 241100, 241130, 241140, 241145, 241390, 241400, 241405, 241410, 241470, 241480, 241485, 241560, 241565, 241610, 241650, 241660, 241760, 241770],
    "Agreste Litoral Sul": [240070, 240180, 240190, 240330, 240540, 240700, 241060, 241110, 241115, 241120, 241160, 241165, 241230, 241240, 241430, 241440, 241520, 241525, 241540, 241545, 241670, 241740, 241750, 241755],
    "Mato Grande": [240200, 240290, 240300, 241080, 241085, 241088, 241200, 241380, 241385, 241388, 241395, 241490, 241530, 241535, 241640, 241645, 241735, 241745, 241765],
    "Assú-Mossoró": [240060, 240065, 240067, 240210, 240340, 240710, 240950, 240960, 241150, 241415, 241420, 241500, 241505, 241700],
    "Potengi": [240220, 240230, 240970, 241075, 241455, 241460, 241510, 241515, 241518, 241519, 241521],
    "Trairi": [240240, 240280, 240285, 241070, 241072, 241095, 241155, 241380, 241446, 241550, 241555, 241556, 241557, 241558, 241720],
    "Sertão Central Cabugi e Litoral Norte": [240080, 240090, 240320, 240580, 240720, 240730, 241105, 241210, 241445, 241450],
    "Sertão do Apodi": [240130, 240140, 240360, 240590, 240740, 240960, 241065, 241220, 241340, 241375, 241376, 241460, 241465, 241550, 241710, 241770, 241780],
    "Terras Potiguares": [240570, 241180, 241250, 241430, 241555]
}

def get_territory(ibge):
    ibge_int = int(str(ibge)[:6])
    for territory, codes in TERRITORIOS.items():
        if ibge_int in codes:
            return territory
    return "Outros/Não mapeado"

def format_excel_sheet(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    navy = "0D1B2A"
    white = "FFFFFF"
    header_font = Font(name='Calibri', bold=True, color=white)
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    
    for col_num, value in enumerate(df.columns.values):
        cell = worksheet.cell(row=1, column=col_num + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, col in enumerate(df.columns):
        ws_col = worksheet.column_dimensions[get_column_letter(i + 1)]
        ws_col.width = 25

def generate_rn_study():
    print("Iniciando Diagnóstico RN PAA (Alinhamento com Estudos)...")
    
    fields = [
        "codigo_ibge", "anomes_s", 
        "agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f",
        "paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i"
    ]
    
    url = f"{MDS_API_URL}?q=codigo_ibge:24*&rows=2000000&wt=csv&fl={','.join(fields)}"
    resp = requests.get(url, verify=False, timeout=120)
    df = pd.read_csv(io.StringIO(resp.text))
    
    df['Ano'] = df['anomes_s'].astype(str).str[:4]
    df['Território'] = df['codigo_ibge'].apply(get_territory)
    
    # Limpeza e Tipagem
    cols_to_sum = ["agricultores_fornec_paa_i", "recur_pagos_agricul_paa_f", "paa_qtd_agricul_familiar_sexo_masculino_i", "paa_qtd_agricul_familiar_sexo_feminino_i"]
    for c in cols_to_sum: df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    
    # 1. RN Geral por Ano (Série Histórica)
    rn_geral = df.groupby('Ano').agg({
        "recur_pagos_agricul_paa_f": "sum",
        "agricultores_fornec_paa_i": "sum",
        "paa_qtd_agricul_familiar_sexo_feminino_i": "sum"
    }).reset_index()
    
    # Cálculo de Capilaridade (Quantos dos 167 municípios receberam recurso no ano)
    capilaridade = df[df['recur_pagos_agricul_paa_f'] > 0].groupby('Ano')['codigo_ibge'].nunique().reset_index()
    capilaridade.columns = ['Ano', 'Municípios com PAA']
    capilaridade['% Capilaridade (RN=167)'] = (capilaridade['Municípios com PAA'] / 167) * 100
    
    rn_geral = pd.merge(rn_geral, capilaridade, on='Ano', how='left')
    rn_geral['% Mulheres no PAA'] = (rn_geral['paa_qtd_agricul_familiar_sexo_feminino_i'] / rn_geral['agricultores_fornec_paa_i']) * 100
    
    # 2. Diagnóstico por Território (Foco do Estudo DIEESE)
    territorial = df.groupby(['Território', 'Ano']).agg({
        "recur_pagos_agricul_paa_f": "sum",
        "agricultores_fornec_paa_i": "sum"
    }).reset_index()
    
    # Capilaridade por Território
    total_munis_territorio = {}
    for t, codes in TERRITORIOS.items(): total_munis_territorio[t] = len(codes)
    
    cap_territorio = df[df['recur_pagos_agricul_paa_f'] > 0].groupby(['Território', 'Ano'])['codigo_ibge'].nunique().reset_index()
    cap_territorio.columns = ['Território', 'Ano', 'Municípios Atendidos']
    
    territorial = pd.merge(territorial, cap_territorio, on=['Território', 'Ano'], how='left')
    territorial['Total Municípios'] = territorial['Território'].map(total_munis_territorio)
    territorial['% Atendimento Territorial'] = (territorial['Municípios Atendidos'] / territorial['Total Municípios']) * 100

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        rn_geral.to_excel(writer, sheet_name="RAIO-X RN (ANUAL)", index=False)
        format_excel_sheet(writer, "RAIO-X RN (ANUAL)", rn_geral)
        
        territorial.to_excel(writer, sheet_name="DIAGNÓSTICO TERRITORIAL", index=False)
        format_excel_sheet(writer, "DIAGNÓSTICO TERRITORIAL", territorial)

    print(f"✅ Diagnóstico RN gerado: {OUTPUT_FILE}")

if __name__ == "__main__":
    generate_rn_study()
